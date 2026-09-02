"""California Contractors State License Board.

The primary discovery source, and the reason the scoring engine stopped guessing.

Every row is a filed record rather than marketing copy, which changes what three
factors can honestly claim:

| CSLB field                | Feeds           | Replaces |
|---------------------------|-----------------|----------|
| `BusinessType`            | succession      | regex for "family owned" on a homepage |
| `IssueDate`               | succession      | regex for "serving since 1985" |
| `WorkersCompCoverageType` | buy_box         | nothing — there was no size signal at all |
| `Classification(s)`       | fragmentation   | fuzzy industry matching |
| `Address`/`City`/`ZIP`    | geography       | map coordinates |

Measured across 12,065 Los Angeles County licences: ownership form and issue
date are present on 100% of rows, against 12% and 10% respectively when the same
signals were scraped from company websites.

**On the workers' compensation field.** California requires a workers' comp
policy of any licensee with employees, and grants an exemption to those without.
So the field separates owner-only operations (5,340 of the 12,065) from staffed
ones (6,684). That is a band, not a headcount, and the factor says so — it can
rule out the zero-employee case, which is definitively below a 10-100 buy box,
but it cannot tell 12 employees from 90.

**What CSLB does not have: websites or email.** Email is withheld by statute
(Business & Professions Code §27) and there is no URL field at all. The website
crawler and the OSM source cover that half; see `app.pipeline.sources`.

The bulk download is a manual step by necessity — the portal's WAF rejects the
form POST that produces a filtered export, though the underlying file endpoints
are plain GETs. The exports are committed under `data/raw/` because California
state material is public domain, which makes the collector reproducible from a
clean clone despite the WAF.
"""

from __future__ import annotations

import csv
import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.pipeline.markets import Market
from app.schemas import Company, Contact, VerificationStatus

logger = logging.getLogger(__name__)

# Trade classifications this buy box covers, mapped onto readable verticals.
CLASSIFICATION_LABELS: dict[str, str] = {
    "C10": "Electrical",
    "C-10": "Electrical",
    "C20": "HVAC",
    "C-20": "HVAC",
    "C36": "Plumbing",
    "C-36": "Plumbing",
}

# Coverage types that mean the licensee employs somebody. "Exempt" is the
# explicit no-employees case; a blank or lapsed policy tells us nothing either
# way and must not be read as either.
_EMPLOYER_COVERAGE = ("workers' compensation", "self-insured", "leasing firm")
_EXEMPT_COVERAGE = "exempt"

_CLASS_SPLIT = re.compile(r"[|,/]")


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _employment(coverage: str | None) -> bool | None:
    """Employees or not, from the workers' comp field. None means unknown."""
    if not coverage:
        return None
    low = coverage.lower()
    if any(marker in low for marker in _EMPLOYER_COVERAGE):
        return True
    if _EXEMPT_COVERAGE in low:
        return False
    # "License does not have current W/C" — lapsed, which says nothing about
    # whether anyone is employed. Unknown, not False.
    return None


def _classifications(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [c.strip().upper() for c in _CLASS_SPLIT.split(raw) if c.strip()]


def _primary_vertical(codes: list[str]) -> str | None:
    """The trade this business is in.

    A licence often carries several classifications — an electrician who also
    holds a plumbing ticket. The buy-box trades win over general ones (a "B"
    general building classification says much less about what the business
    actually does), and the first buy-box code wins among those.
    """
    for code in codes:
        label = CLASSIFICATION_LABELS.get(code)
        if label:
            return label
    return None


def row_to_company(row: dict[str, Any], market: Market) -> Company | None:
    """Map one licence record onto a Company. None if it is unusable."""
    licence = _clean(row.get("LicenseNumber"))
    name = _clean(row.get("BusinessName"))
    if not licence or not name:
        return None

    codes = _classifications(_clean(row.get("Classification(s)")))
    vertical = _primary_vertical(codes)
    if vertical is None:
        # Holds none of the buy-box trades — out of scope rather than low-scoring.
        return None

    issued = _parse_date(row.get("IssueDate"))
    phone = _clean(row.get("PhoneNumber"))
    business_type = _clean(row.get("BusinessType"))
    has_employees = _employment(_clean(row.get("WorkersCompCoverageType")))

    contacts: list[Contact] = []
    if phone:
        contacts.append(
            Contact(
                phone=phone,
                phone_valid=None,  # the validation pass decides
                email=None,  # CSLB never publishes email (B&P Code §27)
                email_status=VerificationStatus.UNKNOWN,
            )
        )

    city = _clean(row.get("City"))
    return Company(
        id=f"cslb:{licence}",
        name=name,
        industry=vertical,
        city=city,
        state=_clean(row.get("State")) or market.state,
        postcode=_clean(row.get("ZIP Code")),
        # Licence issue is a *lower bound* on age — an old firm can hold a newer
        # licence. The succession factor is told which it is.
        founded_year=issued.year if issued else None,
        business_type=business_type,
        has_employees=has_employees,
        licence_number=licence,
        licence_issued=issued,
        licence_classifications=codes,
        contacts=contacts,
        source="cslb",
        source_url="https://www.cslb.ca.gov/OnlineServices/CheckLicenseII/CheckLicense.aspx",
        first_seen=date.today(),
        raw=dict(row),
    )


# Personnel titles, most senior first. A licence often lists one person under
# several titles across time; the strongest is the one worth showing.
_TITLE_RANK = (
    "sole owner",
    "president",
    "chief executive officer",
    "owner",
    "partner",
    "responsible managing owner",
    "responsible managing officer",
    "responsible managing employee",
    "vice president",
    "officer",
    "treasurer",
    "secretary",
)


def _best_title(raw: str) -> str | None:
    """Pick the most senior title from a pipe-separated history."""
    parts = [p.strip() for p in raw.split("|") if p.strip()]
    if not parts:
        return None
    for wanted in _TITLE_RANK:
        for part in parts:
            if wanted in part.lower():
                return part
    return parts[0]


def _person_name(raw: str) -> str | None:
    """CSLB files names as `LAST FIRST MIDDLE`, space-padded. Render them the way
    a person would read them, since this goes in front of a user about to make
    contact."""
    tokens = raw.split()
    if not tokens:
        return None
    if len(tokens) == 1:
        return tokens[0].title()
    last, *rest = tokens
    return " ".join(t.title() for t in [*rest, last])


class CslbSource:
    """Reads committed CSLB exports and yields companies for a market."""

    name = "cslb"

    def __init__(self, raw_dir: Path | None = None) -> None:
        self._raw_dir = raw_dir or Path(__file__).resolve().parents[3].parent / "data" / "raw"

    def load_principals(self) -> dict[str, tuple[str, str | None]]:
        """Licence number -> (person, title), from the committed personnel export.

        **Coverage is partial and the gap is an artefact, not a fact.** The
        statewide personnel file is served through a gateway that closes the
        connection after thirty seconds, which caps any download at roughly 13MB
        however it is requested — plain, gzipped, HTTP/1.0, or from a different
        continent. What arrives is the beginning of a file ordered by licence
        number, so coverage skews toward older licences and stops around 592,000
        of about 1.1 million.

        The consequence matters for how the score reads it: a company with no
        principal here has not been shown to lack one, so contactability reports
        it as a missing signal rather than a finding. Anything else would let a
        download limitation masquerade as a fact about the business.
        """
        path = self._raw_dir / "cslb_la_personnel.csv"
        if not path.exists():
            return {}

        principals: dict[str, tuple[str, str | None]] = {}
        with path.open(encoding="utf-8", newline="") as handle:
            for row in csv.DictReader(handle):
                licence = (row.get("LIC-NO") or "").strip()
                name = _person_name((row.get("Name") or "").strip())
                if not licence or not name or licence in principals:
                    continue
                principals[licence] = (name, _best_title((row.get("EMP-Titl-CDE") or "").strip()))
        return principals

    def _workbooks(self) -> list[Path]:
        if not self._raw_dir.exists():
            return []
        return sorted(self._raw_dir.glob("cslb_*.xlsx"))

    def read_rows(self) -> list[dict[str, Any]]:
        """Every row across every committed export, de-duplicated by licence.

        A contractor holding both an electrical and a plumbing ticket appears in
        two exports; the licence number is the identity, so the second sighting
        is dropped rather than becoming a phantom second business.
        """
        seen: set[str] = set()
        rows: list[dict[str, Any]] = []

        for path in self._workbooks():
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                sheet = workbook.active
                if sheet is None:
                    continue
                iterator = sheet.iter_rows(values_only=True)
                header = [str(h).strip() if h is not None else "" for h in next(iterator, ())]
                count = 0
                for values in iterator:
                    row = dict(zip(header, values, strict=False))
                    licence = _clean(row.get("LicenseNumber"))
                    if not licence or licence in seen:
                        continue
                    seen.add(licence)
                    rows.append(row)
                    count += 1
                logger.info("%s: %d new licences", path.name, count)
            finally:
                workbook.close()

        return rows

    async def discover(self, market: Market) -> list[Company]:
        """Companies for the market. Async only to satisfy the shared interface —
        this source reads local files and never touches the network."""
        if market.state != "CA":
            logger.info(
                "CSLB covers California only; %s is in %s. Returning nothing.",
                market.label,
                market.state,
            )
            return []

        rows = self.read_rows()
        if not rows:
            logger.warning(
                "no CSLB exports found in %s — see the module docstring for how "
                "they are obtained",
                self._raw_dir,
            )
            return []

        principals = self.load_principals()
        wanted_counties = {c.upper() for c in market.counties}
        companies: list[Company] = []
        named = 0

        for row in rows:
            county = (_clean(row.get("County")) or "").upper()
            if wanted_counties and county not in wanted_counties:
                continue
            company = row_to_company(row, market)
            if company is None:
                continue
            if not market.is_core(company.city):
                continue

            principal = principals.get(company.licence_number or "")
            if principal and company.contacts:
                name, title = principal
                company.contacts[0] = company.contacts[0].model_copy(
                    update={"name": name, "title": title, "is_decision_maker": True}
                )
                named += 1

            companies.append(company)

        logger.info(
            "CSLB: %d licences -> %d companies in %s (%d with a named principal)",
            len(rows),
            len(companies),
            market.label,
            named,
        )
        return companies


__all__ = ["CLASSIFICATION_LABELS", "CslbSource", "row_to_company"]
